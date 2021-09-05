from openpyxl import load_workbook
import os

try:
    # Загрузка файла
    wb = load_workbook('mega.xlsx')
    # Получение листа через его имя
    sheet_name = wb.sheetnames[0]
    sheet = wb[sheet_name]
except FileNotFoundError:
    print('Файл "mega.xlsx" не найден\n\nНажмите Enter, чтобы закончить...')
    input()
    exit()

# Словарь для номеров столбцов с данными
params = {
    'Фамилия': None,
    'Имя': None,
    'Отчество': None,
    'Адрес электронной почты': None,
    'Общий стаж': None,
    'Педагогический стаж': None,
    'Сотрудник является классным руководителем в классе': None,
    'Сотрудник является педагогом (Да, Нет)': None,
    'Образование': None,
    'Ученая степень': None,
    'Звание': None,
    'Учёное звание': None,
    'Должность': None,
    'Квалификационная категория': None,
    'ОООД повышения квалификации': None,
    'Название курса': None,
    'Объем курса': None,
    'Дата выдачи': None
}

# Словарь для вывода информации о сотрудниках
teacher = {
    'Фамилия': None,
    'Имя': None,
    'Отчество': None,
    'Адрес электронной почты': None,
    'Общий стаж': None,
    'Педагогический стаж': None,
    'Сотрудник является классным руководителем в классе': None,
    'Сотрудник является педагогом (Да, Нет)': None,
    'Образование': None,
    'Ученая степень': None,
    'Звание': None,
    'Учёное звание': None,
    'Должность': None,
    'Квалификационная категория': None,
    'ОООД повышения квалификации': [],
    'Название курса': [],
    'Объем курса': [],
    'Дата выдачи': []
}

print('Начинаю создание файлов...\n')

try:
    os.mkdir('teachers')
except FileExistsError:
    print('Папка teachers уже существует\n')

# Получение номеров столбцов с необходимыми данными о сотрудиках
col_numb = 1  # Номер столбца
for i in params:  # Шаг по столбцам
    while sheet.cell(2, col_numb + 1).value is not None:
        if sheet.cell(2, col_numb).value == i:
            params[i] = col_numb
            break
        col_numb += 1  # Шаг по строкам

# Заполнение информации об учителях
row_numb = 3  # Номер ряда
while row_numb <= sheet.max_row:
    if sheet.cell(row_numb, 2).value is not None:

        # Получение информации о сотрудниках
        for i in params:  # Шаг по столбцам

            # Получение информации о повышениях квалификации
            if i in ['ОООД повышения квалификации', 'Название курса', 'Объем курса', 'Дата выдачи']:
                if sheet.cell(row_numb, params[i]).value is not None:
                    while sheet.cell(row_numb, params[i]).value is not None and (sheet.cell(row_numb, 2).value == teacher['Фамилия'] or sheet.cell(row_numb, 2).value is None):
                        teacher['ОООД повышения квалификации'].append(sheet.cell(row_numb, params['ОООД повышения квалификации']).value)
                        teacher['Название курса'].append(sheet.cell(row_numb, params['Название курса']).value)
                        teacher['Объем курса'].append(sheet.cell(row_numb, params['Объем курса']).value)
                        teacher['Дата выдачи'].append(sheet.cell(row_numb, params['Дата выдачи']).value.strftime("%Y"))  # Берётся только год
                        row_numb += 1  # Шаг по строкам
                else:
                    teacher[i] = None
                break

            # Получение информации о стаже
            elif i in ['Общий стаж', 'Педагогический стаж']:
                experience = sheet.cell(row_numb, params[i]).value
                for k in experience:

                    # Стаж от пяти лет
                    if k == 'л':
                        teacher[i] = experience[:experience.find("л")] + 'лет'
                        break

                    # Стаж от года до пяти лет
                    elif k == 'г':
                        if int(experience[0]) > 1:
                            teacher[i] = experience[:experience.find("г")] + 'года'  # Стаж больше года, но меньше пяти лет
                        else:
                            teacher[i] = '1 год'  # Стаж в 1 год
                        break

                    # Стаж меньше года
                    elif k in ['м', 'д']:
                        teacher[i] = 'Меньше года'
                        break

            # Получение остальной информации
            else:
                teacher[i] = sheet.cell(row_numb, params[i]).value


        # Вывод информации о преподавателях
        if teacher['Сотрудник является педагогом (Да, Нет)'] == 'Да':
            file_name = teacher['Адрес электронной почты'][:teacher['Адрес электронной почты'].find('@')]
            print(teacher['Фамилия'], end="")
            try:
                os.mkdir('teachers/' + file_name)
            except FileExistsError:
                print(' - папка уже существует, редактирую файл', end="")
            file = open('teachers/' + file_name + '/' + file_name + '.md', 'w')
            file.write('title: \'' + str(teacher['Фамилия']) + ' ' + str(teacher['Имя']) + ' ' + str(teacher['Отчество']) + '\'\n')
            file.write('place_education: \'' + '\'\n')
            file.write('general_experience: \'' + str(teacher['Общий стаж']) + '\'\n')
            file.write('position: \'' + str(teacher['Должность']) + '\'\n')
            file.write('education: \'' + str(teacher['Образование']) + '\'\n')
            file.write('category: \'' + str(teacher['Квалификационная категория']) + '\'\n')
            file.write('experience: \'' + str(teacher['Педагогический стаж']) + '\'\n')
            file.write('email: \'' + str(teacher['Адрес электронной почты']) + '\'\n')
            file.write('course: \n')
            while teacher['ОООД повышения квалификации']:
                file.write('\t-\n')
                file.write('\t\tplace: \'' + str(teacher['ОООД повышения квалификации'][0]) + '\'\n')
                file.write('\t\ttitle: \'' + str(teacher['Название курса'][0]) + '\'\n')
                file.write('\t\thour: \'' + str(teacher['Объем курса'][0]) + '\'\n')
                file.write('\t\tdate: \'' + str(teacher['Дата выдачи'][0]) + '\'\n')
                del teacher['ОООД повышения квалификации'][0]
                del teacher['Название курса'][0]
                del teacher['Объем курса'][0]
                del teacher['Дата выдачи'][0]
            file.write('creator: admin\n')
            file.close
            print(' - готово!\n')

        # Очистка словаря
        for i in params:

            # Очистка словаря от информации о повышениях квалификации
            if i in ['ОООД повышения квалификации', 'Название курса', 'Объем курса', 'Дата выдачи']:
                teacher[i] = []

            # Очистка словаря от остальной информации
            else:
                teacher[i] = None

    row_numb += 1  # Шаг по строкам

print('Всё готово!\n\nНажмите Enter, чтобы закончить...')
input()
