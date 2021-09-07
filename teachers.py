from openpyxl import load_workbook
import os
import re

# Словарь для номеров столбцов с основными данными
teacher_numbers = {
    'Фамилия': None,
    'Имя': None,
    'Отчество': None,
    'Адрес электронной почты': None,
    'Должность': None,
    'Дополнительная должность': None,
    'Группа сотрудников': None,
    'Образование': None,
    'Квалификационная категория по основной должности': None,
    'Общий стаж': None,
    'Педагогический стаж': None,
    'Образовательное учреждение': None,
}

# Словарь для номеров столбцов с данными о курсах повышения квалификации
courses_numbers = {
    'ОООД повышения квалификации (полное наименование)': None,
    'Название курса': None,
    'Объем курса (часы)': None,
    'Дата выдачи': None
}

# Словарь с членами администрации: ключ - фамилия, значение - номер для отображения на сайте
headmaster = {
    'kozyrev': '01.',
    'voznaya': '02.',
    'sidorova': '03.',
    'okova': '04.',
    'pekarskaya': '05.',
    'vishnyakova': '06.',
    'butorina': '07.',
    'aksyutin': '08.',
    'stepanov': '09.'
}

# Словарь для вывода основной информации о сотрудниках
teacher = {
    'Фамилия': None,
    'Имя': None,
    'Отчество': None,
    'Адрес электронной почты': None,
    'Должность': None,
    'Дополнительная должность': None,
    'Группа сотрудников': None,
    'Образование': None,
    'Квалификационная категория по основной должности': None,
    'Общий стаж': None,
    'Педагогический стаж': None,
    'Образовательное учреждение': None,
}
# Словарь для вывода информации о курсах повышения квалификации
courses = {
    'ОООД повышения квалификации (полное наименование)': [],
    'Название курса': [],
    'Объем курса (часы)': [],
    'Дата выдачи': []
}

try:  # Отлов ошибки связанной с отсутствием таблицы с основной информацией
    # Загрузка файла teachers.xlsx
    teacher_wb = load_workbook('teachers.xlsx')
    # Получение листа через его имя
    teacher_sheet_name = teacher_wb.sheetnames[0]
    teacher_sheet = teacher_wb[teacher_sheet_name]
except FileNotFoundError as e:
    print('Ошибка:', e)
    input()
    exit()

# Получение номеров столбцов с необходимыми данными о сотрудиках
col_numb = 1  # Номер столбца
for i in teacher_numbers:  # Шаг по столбцам
    while teacher_sheet.cell(1, col_numb).value is not None:
        if teacher_sheet.cell(1, col_numb).value == i:
            teacher_numbers[i] = col_numb
            break
        col_numb += 1  # Шаг по строкам

try:  # Отлов ошибки связанной с отсутствием таблицы с информацией о курсах повышения квалификации
    # Загрузка файла courses.xlsx
    courses_wb = load_workbook('courses.xlsx')
    # Получение листа через его имя
    courses_sheet_name = courses_wb.sheetnames[0]
    courses_sheet = courses_wb[courses_sheet_name]
except FileNotFoundError as e:
    print('Ошибка:', e)
    input()
    exit()

# Получение номеров столбцов с необходимыми данными о сотрудиках
col_numb = 1  # Номер столбца
for i in courses_numbers:  # Шаг по столбцам
    while courses_sheet.cell(1, col_numb).value is not None:
        if courses_sheet.cell(1, col_numb).value == i:
            courses_numbers[i] = col_numb
            break
        col_numb += 1  # Шаг по строкам

print('Начинаю создание файлов...\n')

# Создание папки teachers
try:
    os.mkdir('teachers')
except FileExistsError:
    print('Папка teachers уже существует\n')

teachers_row_numb = 2  # Номер строки в таблице teachers
courses_row_numb = 2  # Номер строки в таблице courses
number = 1  # Номер учителя без адреса электронной почты
while teachers_row_numb <= teacher_sheet.max_row:
    for i in teacher_numbers:
        # Получение информации о стаже
        if i in ['Общий стаж', 'Педагогический стаж']:
            if i is not None:
                experience = teacher_sheet.cell(teachers_row_numb, teacher_numbers[i]).value
            for k in experience:

                # Стаж не меньше года
                if k in ['л', 'г']:
                    teacher[i] = int(re.findall("\d+", experience)[0])
                    if 5 <= teacher[i] <= 20 or 5 <= teacher[i] % 10 <= 9 or teacher[i] % 10 == 0:
                        teacher[i] = str(teacher[i]) + ' лет'
                    elif 2 <= teacher[i] % 10 <= 4:
                        teacher[i] = str(teacher[i]) + ' года'
                    elif teacher[i] % 10 == 1:
                        teacher[i] = str(teacher[i]) + ' год'
                    break

                # Стаж меньше года
                elif k in ['м', 'д']:
                    teacher[i] = ' меньше года'
                    break
        else:  # Получение остальной информации
            teacher[i] = teacher_sheet.cell(teachers_row_numb, teacher_numbers[i]).value

    # Получение информации о курсах повышения квалификации
    while courses_row_numb <= courses_sheet.max_row and (courses_sheet.cell(courses_row_numb, 2).value == teacher['Фамилия'] or courses_sheet.cell(courses_row_numb, 2).value is None):
        courses['ОООД повышения квалификации (полное наименование)'].append(courses_sheet.cell(courses_row_numb, courses_numbers['ОООД повышения квалификации (полное наименование)']).value)
        courses['Название курса'].append(courses_sheet.cell(courses_row_numb, courses_numbers['Название курса']).value)
        courses['Объем курса (часы)'].append(courses_sheet.cell(courses_row_numb, courses_numbers['Объем курса (часы)']).value)
        courses['Дата выдачи'].append(courses_sheet.cell(courses_row_numb, courses_numbers['Дата выдачи']).value.strftime("%Y"))  # Берётся только год
        courses_row_numb += 1  # Шаг по строкам

    # Создание файла
    if teacher['Адрес электронной почты'] != '' and teacher['Адрес электронной почты'] is not None:
        folder_name = teacher['Адрес электронной почты'][:teacher['Адрес электронной почты'].find('@')]
    else:
        folder_name = 'teacher ' + str(number)
        number += 1
    print(teacher['Фамилия'], end="")

    # Проверка входит ли преподаватель в администрацию
    if folder_name in headmaster.keys():
        folder_name = headmaster[folder_name] + folder_name

    # Проверка наличия папки
    if os.path.exists('teachers_old/' + folder_name + '/teacher.md') or os.path.exists('teachers_old/' + folder_name + '/teacher.en.md'):
        if os.path.exists('teachers_old/' + folder_name + '/teacher.en.md') and not os.path.exists('teachers_old/' + folder_name + '/teacher.md'):
            os.rename('teachers_old/' + folder_name + '/teacher.en.md', 'teachers_old/' + folder_name + '/teacher.md')

        # Открытие файла
        file = open('teachers_old/' + folder_name + '/teacher.md', 'r', encoding="utf-8")

        # taxonomy
        category = ''
        for line in file:
            if 'category' in line:
                category = (re.sub('[^a-z]', '', next(file))).split()[0]
                break

        # class chief
        class_chief = ''
        for line in file:
            if 'class_chief' in line:
                class_chief = line
        file.close()

    # Создание папки
    try:
        os.mkdir('teachers/' + folder_name)
    except FileExistsError as e:
        print('\nОшибка:', e)
        input()
        exit()

    # Открытие файла
    file = open('teachers/' + folder_name + '/teacher.md', 'w', encoding='utf-8')

    # Вывод основных данных об учителе
    file.write('---\n')
    file.write('title: \'' + str(teacher['Фамилия']) + ' ' + str(teacher['Имя']) + ' ' + str(teacher['Отчество']) + '\'\n')
    file.write('place_education: \'' + str(teacher['Образовательное учреждение']) + '\'\n')
    file.write('general_experience: \'' + str(teacher['Общий стаж']) + '\'\n')
    if category != '':
        file.write('taxonomy:\n    category:\n        - ' + category + '\n')
        category = ''
    if class_chief != '':
        file.write(class_chief)
        class_chief = ''
    file.write('position: \'' + str(teacher['Должность']) + '\'\n')
    file.write('education: \'' + str(teacher['Образование']).capitalize() + ' образование\'\n')
    file.write('category: \'' + str(teacher['Квалификационная категория по основной должности']).lower() + '\'\n')
    file.write('experience: \'' + str(teacher['Педагогический стаж']) + '\'\n')
    file.write('email: \'' + str(teacher['Адрес электронной почты']) + '\'\n')
    file.write('course: \n')

    # Вывод данных о курсах повышения квалификации
    while courses['ОООД повышения квалификации (полное наименование)']:
        file.write('    -\n')
        file.write('        place: \'' + str(courses['ОООД повышения квалификации (полное наименование)'][0]) + '\'\n')
        file.write('        title: \'' + str(courses['Название курса'][0]) + '\'\n')
        file.write('        hour: \'' + str(courses['Объем курса (часы)'][0]) + '\'\n')
        file.write('        date: \'' + str(courses['Дата выдачи'][0]) + '\'\n')
        del courses['ОООД повышения квалификации (полное наименование)'][0]
        del courses['Название курса'][0]
        del courses['Объем курса (часы)'][0]
        del courses['Дата выдачи'][0]
    file.write('creator: admin\n')
    file.write('---\n')

    # Закрытие файла
    file.close()
    print(' - готово!\n')

    # Очистка словаря
    for i in teacher:
        teacher[i] = None

    teachers_row_numb += 1

print('Всё готово!\n\nНажмите Enter, чтобы закончить...')
input()
